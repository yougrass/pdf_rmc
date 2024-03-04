import math as math
import numpy as np
from numpy import *
import os,sys
import scipy.integrate as integrate
import scipy.special as special
from matplotlib import pyplot as plt 

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd 
# import xlrd
import random
import datetime
import openpyxl

# Change the common font size
font_size = 14
plt.rcParams.update({'font.size': font_size})




def XYZ_import(file):
    """"
    This function reads an xyz file and imports xyz coordinates
    from the file to the code
    file = the xyz file that you want to use
    """
    
    # the following lines are used for extract data from the file
    xyz_file = open(file)                                    # locally, we use xyz_file represent the file
    data = xyz_file.readlines()                              # data contain all the information of the file
    number_of_atoms = int(data[0])                           # number_of_atoms can help us know how many atoms
    xyz_coordinate_initial = data [2:2+number_of_atoms]      # it will extrac xyz coordinate, initial means we will manipulate 
                                                             #it later
    
    # Now we will get x,y,z coordinate seperately
    x_coordinate = np.zeros((number_of_atoms,1))                             # X_coordinate is x coordinate
    y_coordinate = np.zeros((number_of_atoms,1))                             # Y_coordinate is y coordinate
    z_coordinate = np.zeros((number_of_atoms,1))                             # Z_coordinate is z coordinate
    type_of_atom = []                                        # type_of_atom can know the atom type
    xyz = np.zeros((number_of_atoms,3))                      # we will use xyz-contain 3 component-for further calculation
    
    for i in range (0,number_of_atoms):
        seperate = xyz_coordinate_initial[i].split()         # this will seperate x,y,z coordinate
        x_coordinate[i] = float(seperate[1])
        y_coordinate[i] = float(seperate[2])
        z_coordinate[i] = float(seperate[3])                 # These 3 lines assign the x,y,z seperately
        type_of_atom.insert(i,seperate[0])                   # Type of the atom should assigned indepently
    
    # Now, will need to manipulate the x,y,z coordinate 
    x_coordinate = x_coordinate-x_coordinate.mean()
    y_coordinate = y_coordinate-y_coordinate.mean()
    z_coordinate = z_coordinate-z_coordinate.mean()
    
    for i  in range (0,number_of_atoms):  
        xyz[i]=[x_coordinate[i]]+[y_coordinate[i]]+[z_coordinate[i]]     # This line assigns 3 component to a single variable
    
    XYZ_output = array(xyz)
    xyz_file.close()                                         # File needs to be closed after use
    
    # final return include the xyz coordinate, the atom type and the umber of the atoms we have
    return XYZ_output,type_of_atom,number_of_atoms



def r_space_generation(xyz,length):
    no_change=np.zeros((length-1,length,3))                      # no_change is the matrix that the coordinate did not change
    xyz_change=np.zeros((length-1,length,3))                     # xyz_change is the matrix that coordinate change 
        
    for i in range (0,length-1):
        no_change[i] =xyz.copy()                                 # create the no_change matrix
    
    change_current = list(xyz.copy())                            # current will memory the modification of the coordinate
    for i in range (0,length-1):
        change_current = change_current + [change_current[0]]
        del (change_current[0])
        xyz_change[i] = change_current                           # create the modifided matrix
        
    
    difference = (xyz_change-no_change)*(xyz_change-no_change)   # difference is just the difference of coordinates
    seperate = np.zeros((length-1,length))                       # seperate is the r^2
    for i in range(0,length-1):
        for m in range(0,length):
            seperate[i][m] = sum(difference[i][m])
    
    r = np.sqrt(seperate)                                           # r is a intermediate variable needeed
    r_final = np.zeros((1501,length-1,length))
    for i in range (0,1501):
        r_final[i] = r
        
    return r_final


def factor(atom,length):
    a = openpyxl.load_workbook(r"form_factor.xlsx")
    b = openpyxl.load_workbook(r"python_Final.xlsx")
    sheet_a_1 = a['Sheet1']             # sheet_a_1 will match the element type
    sheet_b_1 = b['Sheet1']             # sheet_b_1 is the factors of the element
    FF = np.zeros((length,1501))
    for i in range (0,length):
        for m in range (1,211):
            if [atom[i]]== sheet_a_1.cell(m+1,1).value.split():
                for zxc in range (0,1501):
                    FF[i][zxc] = sheet_b_1.cell(zxc+2,m+1).value
    return FF


def factor_space_generation (factors,length):
        
    factor_no_change = np.zeros((1501,length-1,length))       # the matrix that did not change
    factor_change =np.zeros((1501,length-1,length))           # the modified changed matrix 
        
    for i in range (0,1501):
        change = list(factors)                                # this is a intermediate valueable to generate the change matrix
        for m in range (0,length-1):
            change = change+ [change[0]]
            del(change[0])
            
            for n in range (0,length):
                factor_no_change[i][m][n] = factors[n][i]
                factor_change[i][m][n] = change[n][i]
    
    factor_space = factor_change*factor_no_change  
    
    return factor_space


def q_space_generation(length):
    
    q_space = np.zeros((1501,length-1,length))                 # this is the final matrix we want to create for q 
    for i in range (0,1501):
        for n in range (0,length-1):
            for m in range (0,length):
                q_space[i][n][m] = (i+100)/100
    
    return q_space



def debye(q,ff,r,TBD,factors):        
    I_Q = np.zeros((1501,))                                                         # this means the final I(Q) we want
    factor_2 = np.zeros((1501,len(factors)))                                        # this used for calculate zero seperation
    for i in range (0,1501):
        for m in range (0,len(factors)):
            factor_2[i][m] = factors[m][i]
            

    for i in range (0,1501):    
        I_Q_1 = factor_2[i]*factor_2[i]*np.exp(-0.5*(TBD*(i+100)/100)**2)          # I_Q_1 is zero seperation part 
        I_Q_2 = sinc(r[i]*q[i]/math.pi)*ff[i]*np.exp(-(TBD*(i+100)/100)**2)        # I_Q_2 is the non-zero seperation part
        I_Q[i] = sum(I_Q_1) + sum(I_Q_2)
        
        
    return I_Q



def correctness():
    correct_parameter = ['Cd','S','C','O','H','In','P','Ni']
    factor_out = factor(correct_parameter,8)
    
    return factor_out



def G_R_generation(I_Q_import,qdamp,Form_F):
    
    # sum_f_q is the sum of f(q)
    sum_f_q = 37*Form_F[0]+20*Form_F[1]+666*Form_F[2]+74*Form_F[3]
    average_f_q = sum_f_q/797

    # sqrsum_f_q is the square addition of all
    sqrsum_f_q = 37*Form_F[0]**2+20*Form_F[1]**2+666*Form_F[2]**2+74*Form_F[3]**2

    # average_square_f_q is the <f(q)^2> average
    average_square_f_q = sqrsum_f_q/797
    
    n1=1
    n2=26
    # Now we need to compute F(Q)
    q_value = np.linspace(n1,n2,1501)
    
    # The next few lines is to find the polynomial fit for Q
    q = np.linspace(n1,n2,1501)
    F_q = (I_Q_import - average_square_f_q)/(average_f_q**2)*average_f_q[0]/1501*q
    S_q = (F_q / q_value) + 1
    
    
    # fig = plt.figure()
    # ax = fig.add_subplot(111)
    # plt.plot(q, I_Q_import, label='I(q)')
    # plt.plot(q,  average_square_f_q, label=r'<ASF$^2$>')
    # # plt.plot(q, average_f_q, label ='?')
    # plt.plot(q, I_Q_import - average_square_f_q, label =r'I(q)-<ASF$^2$>')
    # plt.xlabel(r'$q$ $(\AA ^{-1})$')
    # plt.ylabel(r'$I(q)$')
    # plt.xlim(n1,n2)
    # plt.legend(loc='best')
    # plt.show()

    # # THIS IS THE OLD POLYNOMIAL FIT FOR COMPTON SCATTERING (OUTDATED, NOT WORKING FOR BNL DATA0)
    # n_poly = 0.9 * 16 /math.pi
    # fit1 = np.polyfit(q,F_q,4)
    # fit2 = np.polyfit(q,F_q,5)
    # Fit = np.zeros((1501,))        # Fit is the final polynomial fit
    # for i  in range(0,1501):
    #     p1 = 0 
    #     p2 = 0 
    #     for n in range (0,5):
    #         p1 = p1 + fit1[n]*(((i+100)/100)**(4-n))
    #     for m in range (0,6):
    #         p2 = p2 + fit2[m]*(((i+100)/100)**(5-m))
    #     Fit[i] = (n_poly- 4)*p2+(5-n_poly)*p1
    
    # THIS IS THE NEW, WORKING POLYFIT --> N=7 because Qmax=26 and rpoly=0.85
    fit = np.poly1d(np.polyfit(q, F_q, 7))
    q_space=np.linspace(n1, n2, 1501)
    F_Q_integration = F_q-fit(q_space)
    
    G_R_space = np.zeros((4001,))
    B = np.zeros((4001,))  
    r_value = np.linspace(0,20,4001)
    q_space = np.linspace(n1,n2,1501)

    
    for r in range (0,4001):
        B[r] = np.exp(-(r/100*qdamp)**2/2)
        G_R_space[r] = 2/math.pi*sum(sin(q_space*r_value[r])*F_Q_integration)*B[r]
        
## Uncomment to show nice 2x2 plot of conversion
#     fig, axs = plt.subplots(2, 2)
#     plt.gcf().set_size_inches(12, 6.75)
#     fig.tight_layout()
#     # First, plotting I(q)
#     axs[0,0].plot(q, I_Q_import, label='I(q)')
#     # axs[0,0].plot(q,  average_square_f_q, label=r'<ASF$^2$>')
#     # axs[0,0].plot(q, I_Q_import - average_square_f_q, label =r'I(q)-<ASF$^2$>')
#     axs[0,0].set_xlabel(r'$q$ $(\AA ^{-1})$')
#     axs[0,0].set_ylabel(r'$I(q)$ (a.u.)')
#     axs[0,0].set_xlim(n1,n2)
#     axs[0,0].legend(loc='best')
#     # axs[0,0].set_title('I(q)')
#     # Now, plotting F(q) with polyfit
#     axs[0,1].plot(q,F_q, label='reduced structure factor')
#     axs[0,1].plot(q_space,fit(q_space), label='fit', ls=':')
#     axs[0,1].set_xlabel(r'$q$ $(\AA ^{-1})$')
#     axs[0,1].set_ylabel(r'$F(q)$')
#     axs[0,1].set_xlim(n1,n2)
#     axs[0,1].legend(loc='best')
#     # axs[0,1].set_title('F(q)')
#     # Now, plotting F(q) with polyfit subtracted
#     axs[1,0].plot(q,F_Q_integration,label='F(q), fit subtracted')
#     axs[1,0].set_xlabel(r'$q$ $(\AA ^{-1})$')
#     axs[1,0].set_ylabel(r'$F(q)$ [fit subtracted]')
#     axs[1,0].set_xlim(n1,n2)
#     axs[1,0].legend(loc='best')
#     # axs[1,0].set_title('F(q) [fit subtracted]')
#     # Now, plotting G(r)
#     axs[1,1].plot(r_value, G_R_space, label='G(r)', color='C4')
#     axs[1,1].set_xlabel(r'Inter-atomic spacing $(\AA)$')
#     axs[1,1].set_ylabel(r'$G(r)$')
#     axs[1,1].set_xlim(0,20)
#     axs[1,1].legend(loc='best')
#     # axs[1,1].set_title('G(r)')
#     # fig.suptitle()
#     fig.tight_layout()


    # plt.plot(q,F_q, label='reduced structure factor')
    # plt.plot(q_space,fit(q_space), label='fit', ls=':')
    # plt.xlabel(r'$q$ $(\AA ^{-1})$')
    # plt.ylabel(r'$F(q)$')
    # plt.xlim(n1,n2)
    # plt.legend(loc='best')
    # plt.show()
    
    # # plt.plot(q,average_f_q, label = 'average_f_q')
    # # plt.plot(q,average_square_f_q, label = 'average__square_f_q')
    # # plt.plot(q,S_q, label = 'structure factor')
    # plt.plot(q,F_Q_integration,label='reduced structure factor, fit subtracted')
    # plt.xlabel(r'$q$ $(\AA ^{-1})$')
    # plt.ylabel(r'$F(q)$ [fit subtracted]')
    # plt.xlim(n1,n2)
    # plt.legend(loc='best')
    # # plt.ylim(-0.02,0.02) 
    # plt.show() 
    return G_R_space, B, F_Q_integration


